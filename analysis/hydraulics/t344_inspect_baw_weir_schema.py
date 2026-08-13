"""Read-only schema inspection for the frozen T344 source workbooks.

This script intentionally emits only workbook structure, headers and a few leading
rows.  It does not transform or save a workbook and it does not calculate any T344
endpoint.
"""

from __future__ import annotations

import json
import os
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent / "source_baw_weir"


def scalar(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def inspect(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=False)
    output = {"file": path.name, "sheets": []}
    for sheet in workbook.worksheets:
        leading_rows = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            leading_rows.append([scalar(value) for value in row[:20]])
            if row_index >= 8:
                break
        output["sheets"].append(
            {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "leading_rows": leading_rows,
            }
        )
    workbook.close()
    return output


if __name__ == "__main__":
    representation = os.environ.get("T344_REPRESENTATION", "lab").strip().lower()
    if representation not in {"lab", "num"}:
        raise ValueError("T344_REPRESENTATION must be 'lab' or 'num'")
    files = [
        ROOT / f"Spheres_{representation}_low.xlsx",
        ROOT / f"Spheres_{representation}_medium.xlsx",
        ROOT / f"Spheres_{representation}_high.xlsx",
    ]
    print(json.dumps([inspect(path) for path in files], indent=2))
